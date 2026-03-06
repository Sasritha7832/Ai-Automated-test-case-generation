"""
export_engine.py
----------------
Multi-format QA test suite exporter.

Supported formats:
  1. Excel QA Template  (.xlsx) — formatted like a professional QA management sheet
  2. TestRail CSV       (.csv)  — compatible with TestRail bulk import
  3. Standard CSV       (.csv)  — flat export for Jira/generic QA tools
"""

import io
import csv
import pandas as pd
from typing import List, Dict, Any
from logger import get_logger

logger = get_logger(__name__)

# ─── Color palette (openpyxl PatternFill / Font hex strings) ─────────────────
_COL_HEADER_BG  = "0F3460"   # Dark navy header    (white text)
_COL_META_BG    = "1A2744"   # Meta row bg         (light text)
_COL_PRECOND_BG = "2A2A14"   # Preconditions tint  (yellow text)
_COL_STEP_HDR   = "0A2744"   # Step table header   (sky-blue text)
_COL_STEP_ODD   = "16213E"   # Odd row bg
_COL_STEP_EVEN  = "1A2744"   # Even row bg
_COL_P0         = "FF4B4B"   # Red for P0
_COL_P1         = "FFA500"   # Orange for P1
_COL_P2         = "43D97D"   # Green for P2
_COL_P3         = "17A2B8"   # Blue for P3

_PRIORITY_COLOR = {"P0": _COL_P0, "P1": _COL_P1, "P2": _COL_P2, "P3": _COL_P3}

_STEP_COLS = [
    "Step No", "Description", "Test Data",
    "Expected Result", "Actual Result", "Execution Status",
    "Bug ID", "Notes"
]

# Severity auto-derivation if field is missing
def _derive_severity(test_type: str, priority: str) -> str:
    if priority == "P0":
        return "Critical"
    sev = {
        "Security": "Critical", "Performance": "High", "Negative": "High",
        "Integration": "High",   "API": "High",        "Functional": "Medium",
        "Boundary": "Medium",    "Regression": "Medium", "Data Validation": "Medium",
        "Edge Case": "Medium",   "UI": "Low",           "Accessibility": "Low",
    }
    return sev.get(test_type, "Medium")


def _normalize(tc: dict) -> dict:
    """Ensure all 14 required fields exist in a test case dict."""
    steps = tc.get("steps", tc.get("test_steps", []))
    if isinstance(steps, str):
        steps = [{"step_no": i+1, "description": s.strip(), "test_data": "", "expected_result": ""} for i, s in enumerate(steps.replace(";", "\n").split("\n")) if s.strip()]
    elif isinstance(steps, list):
        for i, s in enumerate(steps):
            if isinstance(s, str):
                steps[i] = {"step_no": i+1, "description": s, "test_data": "", "expected_result": ""}
    elif not isinstance(steps, list):
        steps = [{"step_no": 1, "description": str(steps), "test_data": "", "expected_result": ""}]

    priority  = str(tc.get("priority", "P2"))
    test_type = str(tc.get("test_type", "Functional"))
    return {
        "test_case_id":    tc.get("test_case_id", "TC???"),
        "title":           tc.get("title", tc.get("scenario", "Untitled")),
        "module":          tc.get("module", "General"),
        "test_type":       test_type,
        "scenario":        tc.get("title", tc.get("scenario", "")),
        "preconditions":   tc.get("preconditions", ""),
        "test_steps":      steps,
        "test_data":       tc.get("test_data", ""),
        "expected_result": tc.get("expected_result", ""),
        "actual_result":   tc.get("actual_result", ""),
        "priority":        priority,
        "severity":        tc.get("severity") or _derive_severity(test_type, priority),
        "status":          tc.get("status", "Not Executed"),
        "bug_id":          tc.get("bug_id", ""),
        "notes":           tc.get("notes", "")
    }


# ═══════════════════════════════════════════════════════════════════════════════
# 1. EXCEL QA TEMPLATE
# ═══════════════════════════════════════════════════════════════════════════════

def to_excel_qa_template(
    test_cases: List[Dict[str, Any]],
    project_name: str = "QA Automation Platform",
    author: str = "AI QA Agent",
    reviewer: str = "QA Lead",
    environment: str = "Test / QA Environment",
) -> bytes:
    """
    Returns bytes of a formatted .xlsx file structured like a professional QA template.
    Each test case occupies its own block with header, metadata grid, preconditions,
    and a step execution table.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import (
            Font, PatternFill, Alignment, Border, Side, numbers
        )
        from openpyxl.utils import get_column_letter
    except ImportError:
        logger.error("openpyxl not installed. Excel export unavailable.")
        raise

    logger.info(f"Generating Excel QA template for {len(test_cases)} test cases...")

    wb = Workbook()
    ws = wb.active
    ws.title = "QA Test Suite"

    # ── Column widths ──────────────────────────────────────────────────────────
    col_widths = [6, 40, 25, 40, 25, 18, 12, 20]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ── Helper styles ──────────────────────────────────────────────────────────
    thin = Side(style="thin", color="2A3A5A")
    thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    wrap_align  = Alignment(wrap_text=True, vertical="top")
    centre      = Alignment(horizontal="center", vertical="center", wrap_text=True)

    def _fill(hex_color: str) -> PatternFill:
        return PatternFill("solid", fgColor=hex_color)

    def _font(hex_color: str = "FFFFFF", bold: bool = False, size: int = 10) -> Font:
        return Font(color=hex_color, bold=bold, size=size, name="Calibri")

    def _write(row: int, col: int, value: str,
               fill_hex: str = None, font: Font = None,
               align: Alignment = None, border: bool = True):
        cell = ws.cell(row=row, column=col, value=value)
        if fill_hex:
            cell.fill = _fill(fill_hex)
        if font:
            cell.font = font
        if align:
            cell.alignment = align
        if border:
            cell.border = thin_border
        return cell

    # ── Run: write each test case block ───────────────────────────────────────
    current_row = 1
    NUM_COLS = 8

    # Sort test cases by priority before exporting
    sorted_tcs = sorted(
        test_cases, 
        key=lambda x: str(x.get("priority", "P3") if isinstance(x, dict) else "P3")
    )

    for idx, raw_tc in enumerate(sorted_tcs):
        tc = _normalize(raw_tc)
        steps = tc["test_steps"]
        priority = tc["priority"]
        pri_color = _PRIORITY_COLOR.get(priority, "888888")

        # ── TC TITLE BAR ──────────────────────────────────────────────────────
        ws.merge_cells(
            start_row=current_row, start_column=1,
            end_row=current_row,   end_column=NUM_COLS
        )
        title_cell = ws.cell(
            row=current_row, column=1,
            value=f"  {tc['test_case_id']}  |  {tc['scenario']}"
        )
        title_cell.fill  = _fill(_COL_HEADER_BG)
        title_cell.font  = _font("00D4FF", bold=True, size=11)
        title_cell.alignment = Alignment(vertical="center", wrap_text=True)
        title_cell.border = thin_border
        ws.row_dimensions[current_row].height = 22
        current_row += 1

        # ── METADATA GRID ROW 1 ───────────────────────────────────────────────
        meta1 = [
            ("Project",    project_name),
            ("Module",     tc["module"]),
            ("Test Type",  tc["test_type"]),
            ("Priority",   priority),
            ("Severity",   tc["severity"]),
            ("Status",     tc["status"]),
            ("Author",     author),
            ("Reviewer",   reviewer),
        ]
        # Labels row
        for col_offset, (label, _) in enumerate(meta1, 1):
            _write(current_row, col_offset, label,
                   fill_hex=_COL_STEP_HDR,
                   font=_font("7FC3E8", bold=True, size=9),
                   align=centre)
        current_row += 1
        # Values row
        for col_offset, (_, val) in enumerate(meta1, 1):
            cell = _write(current_row, col_offset, val,
                          fill_hex=_COL_META_BG,
                          font=_font("C9D6E3", size=10),
                          align=centre)
            # Color-code priority cell
            if col_offset == 4:
                cell.font = Font(color=pri_color, bold=True, size=10, name="Calibri")
        ws.row_dimensions[current_row].height = 18
        current_row += 1

        # ── METADATA GRID ROW 2 ───────────────────────────────────────────────
        meta2_labels = ["Environment", "Test Case ID", "Created", "Bug ID", "", "", "", ""]
        meta2_values = [environment, tc["test_case_id"], "Auto-Generated", "—", "", "", "", ""]
        for col_offset, label in enumerate(meta2_labels, 1):
            _write(current_row, col_offset, label,
                   fill_hex=_COL_STEP_HDR,
                   font=_font("7FC3E8", bold=True, size=9),
                   align=centre)
        current_row += 1
        for col_offset, val in enumerate(meta2_values, 1):
            _write(current_row, col_offset, val,
                   fill_hex=_COL_META_BG,
                   font=_font("C9D6E3", size=10),
                   align=centre)
        ws.row_dimensions[current_row].height = 16
        current_row += 1

        # ── PRECONDITIONS ─────────────────────────────────────────────────────
        ws.merge_cells(
            start_row=current_row, start_column=1,
            end_row=current_row,   end_column=NUM_COLS
        )
        prec_cell = ws.cell(
            row=current_row, column=1,
            value=f"  ⚠ PRECONDITIONS: {tc['preconditions'] or 'None specified'}"
        )
        prec_cell.fill      = _fill(_COL_PRECOND_BG)
        prec_cell.font      = _font("FFC107", bold=False, size=10)
        prec_cell.alignment = Alignment(wrap_text=True, vertical="top")
        prec_cell.border    = thin_border
        ws.row_dimensions[current_row].height = 18
        current_row += 1

        # ── STEP TABLE HEADER ─────────────────────────────────────────────────
        for col_offset, hdr in enumerate(_STEP_COLS, 1):
            _write(current_row, col_offset, hdr,
                   fill_hex=_COL_STEP_HDR,
                   font=_font("7FC3E8", bold=True, size=9),
                   align=centre)
        ws.row_dimensions[current_row].height = 18
        current_row += 1

        # ── STEP ROWS ─────────────────────────────────────────────────────────
        test_data   = tc["test_data"]
        expected    = tc["expected_result"]
        step_bg_cycle = [_COL_STEP_ODD, _COL_STEP_EVEN]

        for step_idx, step_obj in enumerate(steps, 1):
            bg = step_bg_cycle[(step_idx - 1) % 2]
            if isinstance(step_obj, dict):
                step_no  = f"{step_obj.get('step_no', step_idx):02d}"
                step_text = step_obj.get("description", str(step_obj))
                td_val   = step_obj.get("test_data", "")
                exp_val  = step_obj.get("expected_result", "")
            else:
                step_no  = f"{step_idx:02d}"
                step_text = str(step_obj)
                td_val = ""
                exp_val = ""

            if not td_val and step_idx == 1:
                td_val = test_data
            if not exp_val and step_idx == len(steps):
                exp_val = expected

            actual_val = tc.get("actual_result", "") or "—"
            bug_val = tc.get("bug_id", "") or "—"
            notes_val = tc.get("notes", "")
            
            row_data = [step_no, step_text, td_val, exp_val, actual_val, "Not Executed", bug_val, notes_val]
            for col_offset, cell_val in enumerate(row_data, 1):
                cell = _write(current_row, col_offset, cell_val,
                              fill_hex=bg,
                              font=_font("C9D6E3", size=10),
                              align=wrap_align)
                if col_offset == 1:  # Step No — center + monospace
                    cell.font = Font(color="00D4FF", bold=True, size=10,
                                     name="Courier New")
                    cell.alignment = centre
                if col_offset == 4:  # Expected Result — green
                    cell.font = _font("C3E6CB", size=10)
                if col_offset == 3:  # Test Data — teal
                    cell.font = _font("98E5BE", size=10)
            ws.row_dimensions[current_row].height = 22
            current_row += 1

        # ── SPACER ROW ────────────────────────────────────────────────────────
        current_row += 1

    # ── Write workbook to bytes ────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    data = buf.read()
    logger.info(f"Excel export generated: {len(data):,} bytes, {len(test_cases)} test cases.")
    return data


# ═══════════════════════════════════════════════════════════════════════════════
# 2. TESTRAIL CSV EXPORT
# ═══════════════════════════════════════════════════════════════════════════════

def to_testrail_csv(test_cases: List[Dict[str, Any]]) -> bytes:
    """
    Returns bytes of a TestRail-compatible CSV for bulk import.
    Required columns: Title, Section, Type, Priority, Preconditions, Steps, Expected Result
    """
    logger.info(f"Generating TestRail CSV for {len(test_cases)} test cases...")

    _TESTRAIL_COLS = [
        "Title", "Section", "Type", "Priority",
        "Preconditions", "Steps", "Expected Result"
    ]

    rows = []
    for raw_tc in test_cases:
        tc = _normalize(raw_tc)
        steps = tc["test_steps"]
        steps_text = "\n".join(
            f"{s.get('step_no', i)}. {s.get('description', str(s))}" for i, s in enumerate(steps, 1)
        )
        rows.append({
            "Title":           tc["scenario"],
            "Section":         tc["module"],
            "Type":            tc["test_type"],
            "Priority":        tc["priority"],
            "Preconditions":   tc["preconditions"],
            "Steps":           steps_text,
            "Expected Result": tc["expected_result"],
        })

    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=_TESTRAIL_COLS, lineterminator="\n")
    writer.writeheader()
    writer.writerows(rows)
    data = buf.getvalue().encode("utf-8")
    logger.info(f"TestRail CSV generated: {len(data):,} bytes.")
    return data


# ═══════════════════════════════════════════════════════════════════════════════
# 3. STANDARD CSV EXPORT
# ═══════════════════════════════════════════════════════════════════════════════

def to_standard_csv(test_cases: List[Dict[str, Any]]) -> bytes:
    """
    Returns bytes of a flat standard CSV for Jira/generic QA tool import.
    Columns: TC ID, Module, Test Type, Scenario, Test Steps, Test Data,
             Expected Result, Actual Result, Execution Status, Bug ID, Notes
    """
    logger.info(f"Generating standard CSV for {len(test_cases)} test cases...")

    _STD_COLS = [
        "TC ID", "Module", "Test Type", "Scenario",
        "Test Steps", "Test Data", "Expected Result",
        "Actual Result", "Execution Status", "Bug ID", "Notes"
    ]

    rows = []
    
    # Sort test cases by priority before exporting
    sorted_tcs = sorted(
        test_cases, 
        key=lambda x: str(x.get("priority", "P3") if isinstance(x, dict) else "P3")
    )

    for raw_tc in sorted_tcs:
        tc = _normalize(raw_tc)
        steps = tc["test_steps"]
        steps_strs = [
            s.get("description", str(s)) if isinstance(s, dict) else str(s)
            for s in steps
        ]
        steps_str = " | ".join(steps_strs)
        rows.append({
            "TC ID":            tc["test_case_id"],
            "Module":           tc["module"],
            "Test Type":        tc["test_type"],
            "Scenario":         tc["scenario"],
            "Test Steps":       steps_str,
            "Test Data":        tc["test_data"],
            "Expected Result":  tc["expected_result"],
            "Actual Result":    tc["actual_result"] or "",
            "Execution Status": tc["status"],
            "Bug ID":           tc["bug_id"] or "",
            "Notes":            tc["notes"] or "",
        })

    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=_STD_COLS, lineterminator="\n")
    writer.writeheader()
    writer.writerows(rows)
    data = buf.getvalue().encode("utf-8")
    logger.info(f"Standard CSV generated: {len(data):,} bytes.")
    return data

